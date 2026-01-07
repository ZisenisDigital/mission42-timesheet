"""
Monthly Timesheet Exporters

Exports monthly timesheets in multiple formats:
- HTML (matching existing timesheet layout)
- CSV (comma-separated values)
- Excel (XLSX format)
"""

import os
import tempfile
from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple
from pathlib import Path

from app.pocketbase_client import PocketBaseClient
from app.config import Config


class MonthlyExporter:
    """
    Exports monthly timesheets in various formats.

    Format matches existing timesheet layout:
    - Title: Zeiterfassung - [Name]
    - Month header with total hours
    - Sequential numbering (0001, 0002, ...)
    - Date format: DD.MM.YYYY
    - Hours in 0.5 increments
    - Location: Remote
    """

    def __init__(self, pb_client: PocketBaseClient, config: Config):
        """
        Initialize monthly exporter.

        Args:
            pb_client: PocketBase client instance
            config: Application configuration
        """
        self.pb_client = pb_client
        self.config = config

    def _get_month_blocks(self, year: int, month: int) -> List[Dict[str, Any]]:
        """
        Fetch all time blocks for a specific month.

        Args:
            year: Year
            month: Month (1-12)

        Returns:
            List of time block dictionaries sorted by date
        """
        # Get start and end of month
        start_date = datetime(year, month, 1)

        # Get end of month
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)

        # Fetch time blocks
        filter_str = (
            f'block_start >= "{start_date.isoformat()}" && '
            f'block_start <= "{end_date.replace(hour=23, minute=59, second=59).isoformat()}"'
        )

        time_blocks = self.pb_client.get_full_list(
            self.pb_client.COLLECTION_TIME_BLOCKS,
            filter=filter_str,
            sort="+block_start"
        )

        # Convert to dict list
        blocks_list = []
        for block in time_blocks:
            if hasattr(block, "__dict__"):
                block_dict = {
                    k: v for k, v in block.__dict__.items()
                    if not k.startswith("_")
                }
            else:
                block_dict = dict(block)
            blocks_list.append(block_dict)

        return blocks_list

    def _format_block_description(self, block: Dict[str, Any]) -> str:
        """
        Format block description for display.

        Args:
            block: Time block dictionary

        Returns:
            Formatted description string
        """
        description = block.get("description", "")

        # Add source context if available
        source = block.get("source", "")
        metadata = block.get("metadata", {})

        # Already formatted descriptions can be returned as-is
        if description:
            return description

        # Fallback formatting based on source
        if source == "wakatime":
            project = metadata.get("project", "Coding")
            return f"Development: {project}"
        elif source == "calendar":
            return "Meeting: " + metadata.get("summary", "Calendar Event")
        elif source == "gmail":
            return "Email: " + metadata.get("subject", "Email")
        elif source == "github":
            return "Development: GitHub"
        elif source == "cloud_events":
            return "Development: Claude Code"
        elif source == "auto_fill":
            return "Development: General"
        else:
            return "Development: Other"

    def export_html(self, year: int, month: int) -> str:
        """
        Export timesheet as HTML.

        Args:
            year: Year
            month: Month (1-12)

        Returns:
            HTML string
        """
        blocks = self._get_month_blocks(year, month)

        # Get export name from settings
        try:
            export_name = self.config.settings.export.export_title_name
        except:
            export_name = "Koni"

        # Calculate total hours
        total_hours = sum(float(block.get("duration_hours", 0)) for block in blocks)

        # Get month name
        month_names = [
            "Januar", "Februar", "März", "April", "Mai", "Juni",
            "Juli", "August", "September", "Oktober", "November", "Dezember"
        ]
        month_name = month_names[month - 1]

        # Generate HTML
        html = f"""<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Zeiterfassung - {export_name} - {month_name} {year}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            font-size: 12px;
        }}
        .header {{
            margin-bottom: 20px;
        }}
        .header h1 {{
            font-size: 24px;
            margin: 0 0 10px 0;
        }}
        .header .info {{
            color: #666;
            margin: 5px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }}
        th {{
            background-color: #f2f2f2;
            font-weight: bold;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .total-row {{
            font-weight: bold;
            background-color: #e8e8e8 !important;
        }}
        .number {{
            text-align: right;
            font-family: monospace;
        }}
        .hours {{
            text-align: right;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Zeiterfassung - {export_name}</h1>
        <div class="info">
            <strong>Monat:</strong> {month_name} {year}
        </div>
        <div class="info">
            <strong>Erstellt am:</strong> {datetime.now().strftime("%d.%m.%Y %H:%M")} |
            <strong>Gesamt:</strong> {total_hours:.1f} Stunden
        </div>
    </div>

    <table>
        <thead>
            <tr>
                <th style="width: 80px;">Nr.</th>
                <th style="width: 120px;">Datum</th>
                <th style="width: 100px;">Stunden</th>
                <th>Beschreibung</th>
                <th style="width: 100px;">Ort</th>
            </tr>
        </thead>
        <tbody>
"""

        # Add rows
        for idx, block in enumerate(blocks, start=1):
            # Parse block_start
            block_start = block.get("block_start", "")
            if isinstance(block_start, str):
                try:
                    block_date = datetime.fromisoformat(block_start.replace("Z", "+00:00"))
                except:
                    block_date = datetime.now()
            else:
                block_date = block_start

            # Format date as DD.MM.YYYY
            date_str = block_date.strftime("%d.%m.%Y")

            # Get hours
            hours = float(block.get("duration_hours", 0))

            # Get description
            description = self._format_block_description(block)

            # Get location (always Remote)
            location = "Remote"

            # Add row
            html += f"""            <tr>
                <td class="number">{idx:04d}</td>
                <td>{date_str}</td>
                <td class="hours">{hours:.1f}</td>
                <td>{description}</td>
                <td>{location}</td>
            </tr>
"""

        # Add total row
        html += f"""            <tr class="total-row">
                <td colspan="2"><strong>Gesamt:</strong></td>
                <td class="hours"><strong>{total_hours:.1f}</strong></td>
                <td colspan="2"></td>
            </tr>
        </tbody>
    </table>
</body>
</html>
"""

        return html

    def export_csv(self, year: int, month: int) -> str:
        """
        Export timesheet as CSV.

        Args:
            year: Year
            month: Month (1-12)

        Returns:
            CSV string
        """
        import csv
        from io import StringIO

        blocks = self._get_month_blocks(year, month)

        # Calculate total hours
        total_hours = sum(float(block.get("duration_hours", 0)) for block in blocks)

        output = StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(["Nr.", "Datum", "Stunden", "Beschreibung", "Ort"])

        # Write data rows
        for idx, block in enumerate(blocks, start=1):
            # Parse block_start
            block_start = block.get("block_start", "")
            if isinstance(block_start, str):
                try:
                    block_date = datetime.fromisoformat(block_start.replace("Z", "+00:00"))
                except:
                    block_date = datetime.now()
            else:
                block_date = block_start

            # Format date as DD.MM.YYYY
            date_str = block_date.strftime("%d.%m.%Y")

            # Get hours
            hours = float(block.get("duration_hours", 0))

            # Get description
            description = self._format_block_description(block)

            # Get location
            location = "Remote"

            writer.writerow([f"{idx:04d}", date_str, f"{hours:.1f}", description, location])

        # Write total row
        writer.writerow(["Gesamt:", "", f"{total_hours:.1f}", "", ""])

        return output.getvalue()

    def export_excel(self, year: int, month: int) -> str:
        """
        Export timesheet as Excel file.

        Args:
            year: Year
            month: Month (1-12)

        Returns:
            Path to temporary Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: uv add openpyxl")

        blocks = self._get_month_blocks(year, month)

        # Get export name from settings
        try:
            export_name = self.config.settings.export.export_title_name
        except:
            export_name = "Koni"

        # Calculate total hours
        total_hours = sum(float(block.get("duration_hours", 0)) for block in blocks)

        # Get month name
        month_names = [
            "Januar", "Februar", "März", "April", "Mai", "Juni",
            "Juli", "August", "September", "Oktober", "November", "Dezember"
        ]
        month_name = month_names[month - 1]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{month_name} {year}"

        # Title
        ws["A1"] = f"Zeiterfassung - {export_name}"
        ws["A1"].font = Font(size=16, bold=True)

        # Info
        ws["A2"] = f"Monat: {month_name} {year}"
        ws["A3"] = f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Gesamt: {total_hours:.1f} Stunden"

        # Headers
        headers = ["Nr.", "Datum", "Stunden", "Beschreibung", "Ort"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Data rows
        row = 6
        for idx, block in enumerate(blocks, start=1):
            # Parse block_start
            block_start = block.get("block_start", "")
            if isinstance(block_start, str):
                try:
                    block_date = datetime.fromisoformat(block_start.replace("Z", "+00:00"))
                except:
                    block_date = datetime.now()
            else:
                block_date = block_start

            # Format date as DD.MM.YYYY
            date_str = block_date.strftime("%d.%m.%Y")

            # Get hours
            hours = float(block.get("duration_hours", 0))

            # Get description
            description = self._format_block_description(block)

            # Get location
            location = "Remote"

            ws.cell(row=row, column=1, value=f"{idx:04d}")
            ws.cell(row=row, column=2, value=date_str)
            ws.cell(row=row, column=3, value=hours)
            ws.cell(row=row, column=4, value=description)
            ws.cell(row=row, column=5, value=location)

            row += 1

        # Total row
        ws.cell(row=row, column=1, value="Gesamt:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_hours)
        ws.cell(row=row, column=3).font = Font(bold=True)

        # Make total row bold
        for col in range(1, 6):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"
            )

        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 12

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        return temp_file.name
